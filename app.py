import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

st.title("JIRA → LDSO Generator 🚀")

jira_file = st.file_uploader("Upload JIRA", type=["csv"])
template_file = st.file_uploader("Upload Template", type=["xlsx"])

if jira_file and template_file:

    df_jira = pd.read_csv(jira_file)

    # ===== SYSTEM LOGIC =====
    def map_system(summary):
        if pd.isna(summary):
            return None
        s = summary.upper()
        if 'AZURE' in s:
            return 'TCAP Cloud'
        elif 'L-DCM' in s:
            return 'LDCM'
        return 'Other'

    df_jira['System'] = df_jira['Summary'].apply(map_system)

    # ===== LOAD TEMPLATE (IMPORTANT) =====
    wb = load_workbook(template_file)
    ws = wb.active

    # ===== EXAMPLE: FILL DATA =====
    # 🔥 ตรงนี้คือ key ของงานนี้ (ต้อง match position)

    # Total Incident
    total_incident = len(df_jira[df_jira['Issue Type'] == 'Incident'])
    ws['C5'] = total_incident

    # Rank A / B / C (สมมติใช้ Priority)
    ws['C7'] = len(df_jira[df_jira['Priority'] == 'Highest'])
    ws['C8'] = len(df_jira[df_jira['Priority'] == 'High'])
    ws['C9'] = len(df_jira[df_jira['Priority'] == 'Medium'])

    # Status count
    ws['F9'] = len(df_jira[df_jira['Status'] == 'Investigating'])
    ws['G9'] = len(df_jira[df_jira['Status'] == 'Fixing'])
    ws['H9'] = len(df_jira[df_jira['Status'] == 'Under Confirmation'])

    # ===== EXPORT =====
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    st.download_button(
        "📥 Download LDSO",
        data=output,
        file_name="LDSO_Output.xlsx"
    )
